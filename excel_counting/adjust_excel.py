from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter


def adjust_excel_column_width(output_file, max_column_width, column_name):
    try:
        workbook = load_workbook(output_file)
        worksheet = workbook.active

        for column in worksheet.iter_cols(min_row=1, max_col=worksheet.max_column, min_col=1, values_only=False):
            if column[0].value == column_name:
                column_index = column[0].column
                column_letter = get_column_letter(column_index)
                worksheet.column_dimensions[column_letter].width = max_column_width
                break
        else:
            print(f"列名 '{column_name}' 在工作表中不存在。")
            return

        workbook.save(output_file)
    except Exception as e:
        print(f"调整Excel列宽时发生错误: {e}")
