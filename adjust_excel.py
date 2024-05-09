from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter


def adjust_excel_column_width(output_file, max_column_width, column_name):
    try:
        # 加载已保存的Excel工作簿
        workbook = load_workbook(output_file)
        worksheet = workbook.active

        # 查找列名对应的列
        for column in worksheet.iter_cols(min_row=1, max_col=worksheet.max_column, min_col=1, values_only=False):
            if column[0].value == column_name:
                # 获取列索引
                column_index = column[0].column
                # 将索引转换为Excel列字母
                column_letter = get_column_letter(column_index)
                # 设置Excel工作表中相应列的宽度
                worksheet.column_dimensions[column_letter].width = max_column_width
                break
        else:
            # 如果找不到列名，打印错误消息
            print(f"列名 '{column_name}' 在工作表中不存在。")
            return

        # 保存修改后的工作簿
        workbook.save(output_file)
    except Exception as e:
        print(f"调整Excel列宽时发生错误: {e}")